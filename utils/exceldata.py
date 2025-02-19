import datetime
import os
import openpyxl

from datetime import datetime
from openpyxl.workbook import Workbook
from utils.testdataset import TestDataSet
from logging import Logger

class ExcelData(TestDataSet):
    """
    Class that provides the methods to handle the test data in Excel files.
    :param test_scenario: The name of the test scenario.
    :param testcase: The name of the current test case.
    :param log: Logger object.
    """

    DIR_PREFIX: str = "../testdata/"
    EXCEL_ARCHIVE: str = "test_archive.xlsx"
    DATASET_TRANSACTION: str = "dataset_transaction.xlsx"
    DATASET_ACCOUNT: str = "dataset_account.xlsx"
    DATASET_CUSTOMER: str = "dataset_customer.xlsx"

    def __init__(self, test_scenario: str, testcase: str, log: Logger):
        super().__init__(test_scenario, testcase)
        self._log: Logger = log

    @staticmethod
    def get_excel_data(file_name: str, log: Logger, sheet_name: str = None) -> list [dict]:
        """
        Returns the data from the Excel file as a dictionary.
        The first row should contain the column names used for the dictionary keys.
        :param file_name: Filename of the Excel file.
        :param log: Logger object used to log the info & errors.
        :param sheet_name:  The name of the sheet in the workbook.
        :return: List of dictionaries w/ data.
        """
        try:
            book = openpyxl.load_workbook(ExcelData.DIR_PREFIX + file_name)
            if sheet_name:
                sheet = book[sheet_name]
            else:
                sheet = book.active
            # Declaring temporary list
            excel_data: list = []
            # Reading the data
            for row in range(2, sheet.max_row+1):
                temp_dict: dict = {}
                for column in range(1, sheet.max_column+1):
                    temp_dict[sheet.cell(row = 1, column = column).value] = sheet.cell(row = row, column = column).value
                excel_data.append(temp_dict)
        except FileNotFoundError:
            log.warning(f"File {ExcelData.DIR_PREFIX + file_name} not found.")
        return excel_data

    @staticmethod
    def save_excel_data(file_name: str, data: list [dict], log: Logger) -> None:
        """
        Saves the formatted data to the Excel file. The first row will contain the
        column names based on the dictionary keys.
        :param file_name: Filename of the Excel file
        :param data: List of dictionaries containing the data to be written. First list should
        should contain the table headers.
        :param log: Logger object used to log the info & errors.
        :return: None.
        """
        book = Workbook()
        sheet = book.active
        sheet.title = "Data"
        # Adding the header w/ column names
        column_names = data[0].keys()
        # Start from the 1 because the cells are indexed from 1
        for key, name in enumerate(column_names, 1):
            sheet.column_dimensions[sheet.cell(row = 1, column = key).column_letter].width = 30
            sheet.cell(row = 1, column = key).value = name
        # Adding the data in rows
        for r_key, row in enumerate(data, 1):
            values: list = list(row.values())
            for c_key, column in enumerate(row, 1):
                sheet.cell(row = r_key + 1, column = c_key).value = values[c_key - 1]
        # Saving the file
        try:
            book.save(ExcelData.DIR_PREFIX + file_name)
        except IOError:
            log.warning(f"Could not open the file {ExcelData.DIR_PREFIX + file_name} for writing.")

    def save_data(self) -> None:
        """
        Saves the data to the Excel file.
        The file name is created based on the test scenario name.
        The data for each test case is written to the sheets with the name of the test case.
        :return: None.
        """
        # Getting the test datetime
        date_time = datetime.now()
        # Loading the existing data or creating the new workbook
        if os.path.exists(self.DIR_PREFIX + self.EXCEL_ARCHIVE):
            try:
                book = openpyxl.load_workbook(self.DIR_PREFIX + self.EXCEL_ARCHIVE)
            except IOError:
                self._log.warning(f"Could not open the file {self.DIR_PREFIX + self.EXCEL_ARCHIVE}.")
        else:
            book = Workbook()
        # Switching to the sheet or creating it
        if self._testcase in book.sheetnames:
            sheet = book[self._testcase]
        else:
            if book.sheetnames[0] == "Sheet":
                sheet = book.active
                sheet.title = self._testcase
            else:
                book.create_sheet(self._testcase)
                sheet = book[self._testcase]
            # Adding the header w/ column names
            sheet["A1"] = self._test_scenario
            sheet.column_dimensions["A"].width = 30
            sheet["B1"] = "Date & time"
            sheet.column_dimensions["B"].width = 30
            column_names = self._data[self._testcase].keys()
            # Start from the number 3 to make room for the test scenario name and date-time
            for key, name in enumerate(column_names, 3):
                sheet.column_dimensions[sheet.cell(row = 1, column = key).column_letter].width = 30
                sheet.cell(row = 1, column = key).value = name
        no_of_rows = sheet.max_row
        if no_of_rows >= 2:
            target_row = no_of_rows + 1
        else:
            target_row = 2
        # Adding the test scenario name and date-time
        sheet.cell(row = target_row, column = 1).value = self._test_scenario
        sheet.cell(row = target_row, column = 2).value = date_time
        # Adding the data in columns
        values: list = list(self._data[self._testcase].values())
        for c_key, column in enumerate(self._data[self._testcase], 3):
            sheet.cell(row = target_row, column = c_key).value = values[c_key - 3]
        # Saving the file
        try:
            book.save(self.DIR_PREFIX + self.EXCEL_ARCHIVE)
        except IOError:
            self._log.warning(f"Could not open the file {self.DIR_PREFIX + self.EXCEL_ARCHIVE} for writing.")

