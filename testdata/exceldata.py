import datetime
import os
from datetime import datetime

import openpyxl
from openpyxl.workbook import Workbook


class ExcelData():

    def __init__(self, testcase: str):
        self._data: dict = {}
        self._testcase: str = testcase

    @staticmethod
    def get_excel_data(file_name: str, sheet_name: str = None) -> list [dict]:
        """
        Returns the data from the Excel file as a dictionary.
        The first row should contain the column names used for the
        dictionary keys.
        :param file_name: Filename of the Excel file.
        :param sheet_name:  Let's you select the sheet in the workbook.
        :return: List of dictionaries w/ data.
        """
        book = openpyxl.load_workbook(f"../testdata/{file_name}")
        if sheet_name:
            sheet = book[sheet_name]
        else:
            sheet = book.active
        # Declaring temporary list
        excel_data: list = []
        # Reading the data
        for row in range(2, sheet.max_row+1):
            temp_dict: dict = {}
            for column in range(1, sheet.max_column+1):
                temp_dict[sheet.cell(row = 1, column = column).value] = sheet.cell(row = row, column = column).value
            excel_data.append(temp_dict)
        return excel_data


    @staticmethod
    def save_excel_data(file_name: str, data: list [dict]) -> None:
        """
        Lets you save the formatted data to the Excel file. The first row
        will contain the column names used for the dictionary keys.
        :param file_name: Filename of the Excel file
        :param data: List of lists containing the data to be written. First list should
        should contain the table headers.
        :return:
        """
        book = Workbook()
        sheet = book.active
        sheet.title = "Data"
        # Adding the header w/ column names
        column_names = data[0].keys()
        # Start from the 1 because the cells are indexed from 1
        for key, name in enumerate(column_names, 1):
            sheet.column_dimensions[sheet.cell(row = 1, column = key).column_letter].width = 30
            sheet.cell(row = 1, column = key).value = name
        # Adding the data in rows
        for r_key, row in enumerate(data, 1):
            values: list = list(row.values())
            for c_key, column in enumerate(row, 1):
                sheet.cell(row = r_key + 1, column = c_key).value = values[c_key - 1]
        # Saving the file
        book.save(f"../testdata/{file_name}")
        # book.save(f"./{file_name}")

    def add_data(self, key: str, value: str) -> None:
        """
        Allows you to add the test case data to the dictionary.
        :param key: The key of the dictionary entry.
        :param value: The value of the dictionary entry.
        :return:
        """
        self._data[key] = value

    def remove_data_by_key(self, key: str) -> None:
        """
        Allows you to remove the test case data from the dictionary by key.
        :param key: The key of the dictionary entry.
        :return:
        """
        self._data.pop(key)

    def remove_data_by_value(self, value: str) -> None:
        """
        Allows you to remove the test case data from the dictionary by value.
        :param value: The value of the dictionary entry.
        :return:
        """
        self._data = { key:val for key, val in self._data.items() if val != value }

    def save_data(self) -> None:
        """
        Saves the data to the Excel file.
        The file name is created based on the date & the test case name.
        :return:
        """
        book = Workbook()
        sheet = book.active
        sheet.title = f"Data: {self._testcase}"
        # Adding the header w/ column names
        column_names = self._data.keys()
        # Start from the number 2 to make room for the test case name
        for key, name in enumerate(column_names, 2):
            sheet.column_dimensions[sheet.cell(row = 1, column = key).column_letter].width = 30
            sheet.cell(row = 1, column = key).value = name
        # Adding the test case name
        sheet.cell(row = 2, column = 1).value = self._testcase
        # Adding the data in columns
        values: list = list(self._data.values())
        for c_key, column in enumerate(self._data, 2):
            sheet.cell(row = 2, column = c_key).value = values[c_key - 1]
        # Saving the file
        date = datetime.now().date()
        time = datetime.now().time().strftime("%H:%M:%S")
        file_name = f"{self._testcase}_{date}_{time}.xlsx"
        book.save(f"../testdata/{file_name}")


    def read_data(self, testcase: str, date: str) -> dict | list [dict]:
        """
        Reads the data from the Excel file or mutiple files if there are multiple test cases
        with the same name & date.
        :param testcase: The name of the testcase for which the data should be read.
        :param date: The date when the testcase was ran.
        :return:
        """
        files = os.listdir("../")
        valid_files = [file for file in files if testcase in file and date in file]
        if len(valid_files) == 1:
            book = openpyxl.load_workbook(f"../testdata/{valid_files[0]}")
            sheet = book.active
            # Reading the data
            for row in range(2, sheet.max_row + 1):
                excel_data: dict = {}
                for column in range(1, sheet.max_column + 1):
                    excel_data[sheet.cell(row=1, column=column).value] = sheet.cell(row=row, column=column).value
        elif len(valid_files) > 1:
            for file in valid_files:
                book = openpyxl.load_workbook(f"../testdata/{file}")
                sheet = book.active
                # Declaring temporary list
                excel_data: list = []
                # Reading the data
                for row in range(2, sheet.max_row + 1):
                    temp_dict: dict = {}
                    for column in range(1, sheet.max_column + 1):
                        temp_dict[sheet.cell(row=1, column=column).value] = sheet.cell(row=row, column=column).value
                    excel_data.append(temp_dict)
        else:
            excel_data = None
        return excel_data
